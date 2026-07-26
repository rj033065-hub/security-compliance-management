import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from services.timezone_service import get_ist_now, format_ist

class ReportGenerator:
    @staticmethod
    def generate_excel_compliance_report(stats, frameworks, tasks, audits, risks):
        wb = Workbook()
        
        # Sheet 1: Executive Summary
        ws1 = wb.active
        ws1.title = "Executive Summary"
        
        # Styles
        header_font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A2530", end_color="1A2530", fill_type="solid")
        title_font = Font(name="Segoe UI", size=16, bold=True, color="0B3C5D")
        sub_font = Font(name="Segoe UI", size=11, bold=True, color="333333")
        
        ws1.append(["SECURITY COMPLIANCE MANAGEMENT SYSTEM"])
        ws1.append(["Executive Compliance Summary Report"])
        ws1.append([f"Generated Date (IST): {format_ist(get_ist_now())}"])
        ws1.append([])
        
        ws1.append(["Metric", "Value"])
        ws1.append(["Total Employees", stats.get('total_employees', 0)])
        ws1.append(["Total Security Policies", stats.get('total_policies', 0)])
        ws1.append(["Active Frameworks", stats.get('total_frameworks', 0)])
        ws1.append(["Overall Compliance Score", f"{stats.get('compliance_score', 0)}%"])
        ws1.append(["Completed Tasks", stats.get('completed_tasks', 0)])
        ws1.append(["Pending / In-Progress Tasks", stats.get('pending_tasks', 0)])
        ws1.append(["High / Critical Risk Items", stats.get('high_risks', 0)])
        ws1.append(["Total Scheduled/Active Audits", stats.get('total_audits', 0)])
        
        # Format table
        ws1.merge_cells("A1:B1")
        ws1.merge_cells("A2:B2")
        ws1.merge_cells("A3:B3")
        ws1["A1"].font = title_font
        ws1["A2"].font = sub_font
        ws1["A3"].font = Font(name="Segoe UI", size=9, italic=True, color="666666")
        
        ws1["A5"].font = Font(bold=True, color="FFFFFF")
        ws1["A5"].fill = header_fill
        ws1["B5"].font = Font(bold=True, color="FFFFFF")
        ws1["B5"].fill = header_fill
        
        # Sheet 2: Frameworks & Controls
        ws2 = wb.create_sheet(title="Frameworks & Controls")
        ws2.append(["Framework Name", "Code", "Version", "Status", "Control Code", "Control Name", "Risk Level", "Implementation Status"])
        
        ws2["A1"].font = Font(bold=True, color="FFFFFF")
        ws2["A1"].fill = header_fill
        for col in ["B1", "C1", "D1", "E1", "F1", "G1", "H1"]:
            ws2[col].font = Font(bold=True, color="FFFFFF")
            ws2[col].fill = header_fill
            
        for fw in frameworks:
            for ctrl in fw.controls:
                ws2.append([
                    fw.name,
                    fw.code,
                    fw.version,
                    fw.status,
                    ctrl.control_code,
                    ctrl.name,
                    ctrl.risk_level,
                    ctrl.status
                ])
                
        # Sheet 3: Tasks & Evidence Tracking
        ws3 = wb.create_sheet(title="Task Management")
        ws3.append(["Task ID", "Title", "Assigned To", "Priority", "Due Date", "Status", "Control Code"])
        for col in ["A1", "B1", "C1", "D1", "E1", "F1", "G1"]:
            ws3[col].font = Font(bold=True, color="FFFFFF")
            ws3[col].fill = header_fill
            
        for t in tasks:
            ws3.append([
                f"TSK-{t.id:04d}",
                t.title,
                t.assignee.username if t.assignee else 'Unassigned',
                t.priority,
                t.due_date.strftime("%Y-%m-%d") if t.due_date else '',
                t.status,
                t.control.control_code if t.control else 'General'
            ])
            
        # Sheet 4: Audits & Findings
        ws4 = wb.create_sheet(title="Audit Log")
        ws4.append(["Audit Code", "Audit Name", "Framework", "Auditor", "Start Date", "End Date", "Status", "Final Status"])
        for col in ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1"]:
            ws4[col].font = Font(bold=True, color="FFFFFF")
            ws4[col].fill = header_fill
            
        for a in audits:
            ws4.append([
                a.audit_code,
                a.name,
                a.framework.name if a.framework else 'N/A',
                a.auditor.username if a.auditor else 'Unassigned',
                a.start_date.strftime("%Y-%m-%d") if a.start_date else '',
                a.end_date.strftime("%Y-%m-%d") if a.end_date else '',
                a.status,
                a.final_status
            ])

        # Auto adjust column widths safely
        for sheet in wb.worksheets:
            for col in sheet.columns:
                col_letter = get_column_letter(col[0].column)
                valid_lengths = []
                for cell in col:
                    if isinstance(cell, MergedCell):
                        continue
                    if sheet.merged_cells:
                        is_in_merged = False
                        for m_range in sheet.merged_cells.ranges:
                            if cell.coordinate in m_range and (m_range.min_col != m_range.max_col):
                                is_in_merged = True
                                break
                        if is_in_merged:
                            continue
                    val_str = str(cell.value or '')
                    if val_str:
                        valid_lengths.append(len(val_str))

                max_len = max(valid_lengths) if valid_lengths else 10
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream

    @staticmethod
    def generate_pdf_report(stats, frameworks, tasks, audits, risks):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=22,
            leading=26,
            textColor=colors.HexColor('#0F172A'),
            alignment=1, # Center
            spaceAfter=10
        )
        
        subtitle_style = ParagraphStyle(
            'DocSubTitle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            leading=14,
            textColor=colors.HexColor('#475569'),
            alignment=1,
            spaceAfter=20
        )

        h2_style = ParagraphStyle(
            'H2Heading',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=14,
            leading=18,
            textColor=colors.HexColor('#1E293B'),
            spaceBefore=15,
            spaceAfter=8
        )
        
        body_style = ParagraphStyle(
            'Body',
            parent=styles['BodyText'],
            fontName='Helvetica',
            fontSize=9,
            leading=12,
            textColor=colors.HexColor('#334155')
        )
        
        story = []
        
        # Document Header
        story.append(Paragraph("SECURITY COMPLIANCE MANAGEMENT SYSTEM", title_style))
        story.append(Paragraph(f"Official Audit & Executive Compliance Assessment Report — Generated (IST): {format_ist(get_ist_now())}", subtitle_style))
        story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#2563EB'), spaceAfter=15))
        
        # Summary KPI Table
        story.append(Paragraph("1. Executive Summary & KPIs", h2_style))
        
        kpi_data = [
            [Paragraph("<b>Metric Name</b>", body_style), Paragraph("<b>Status / Count</b>", body_style), Paragraph("<b>Metric Name</b>", body_style), Paragraph("<b>Status / Count</b>", body_style)],
            ["Total Employees", str(stats.get('total_employees', 0)), "Overall Compliance Score", f"{stats.get('compliance_score', 0)}%"],
            ["Total Policies", str(stats.get('total_policies', 0)), "Completed Tasks", str(stats.get('completed_tasks', 0))],
            ["Active Frameworks", str(stats.get('total_frameworks', 0)), "Pending Tasks", str(stats.get('pending_tasks', 0))],
            ["Audits Conducted", str(stats.get('total_audits', 0)), "High/Critical Risks", str(stats.get('high_risks', 0))]
        ]
        
        t_kpi = Table(kpi_data, colWidths=[150, 100, 150, 100])
        t_kpi.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F1F5F9')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#0F172A')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(t_kpi)
        story.append(Spacer(1, 15))
        
        # Frameworks Overview
        story.append(Paragraph("2. Active Compliance Frameworks & Controls", h2_style))
        fw_data = [[Paragraph("<b>Framework</b>", body_style), Paragraph("<b>Code</b>", body_style), Paragraph("<b>Version</b>", body_style), Paragraph("<b>Status</b>", body_style), Paragraph("<b>Controls Count</b>", body_style)]]
        for fw in frameworks:
            fw_data.append([
                fw.name,
                fw.code,
                fw.version,
                fw.status,
                str(len(fw.controls))
            ])
            
        t_fw = Table(fw_data, colWidths=[160, 90, 70, 80, 100])
        t_fw.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#94A3B8')),
            ('PADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(t_fw)
        story.append(Spacer(1, 15))

        # Recent Audits
        story.append(Paragraph("3. Audits & Compliance Status", h2_style))
        audit_data = [[Paragraph("<b>Audit Code</b>", body_style), Paragraph("<b>Audit Title</b>", body_style), Paragraph("<b>Auditor</b>", body_style), Paragraph("<b>Status</b>", body_style), Paragraph("<b>Result</b>", body_style)]]
        for a in audits[:5]:
            audit_data.append([
                a.audit_code,
                a.name[:35] + ('...' if len(a.name) > 35 else ''),
                a.auditor.username if a.auditor else 'N/A',
                a.status,
                a.final_status
            ])
        if len(audit_data) == 1:
            audit_data.append(["N/A", "No audits registered yet", "N/A", "N/A", "N/A"])
            
        t_audit = Table(audit_data, colWidths=[90, 180, 90, 70, 70])
        t_audit.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F766E')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCFBF1')),
            ('PADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(t_audit)
        story.append(Spacer(1, 20))
        
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#CBD5E1'), spaceAfter=10))
        story.append(Paragraph("<i>Report generated automatically by Security Compliance Management System AI Platform.</i>", ParagraphStyle('Footer', parent=styles['Italic'], fontSize=8, textColor=colors.HexColor('#64748B'), alignment=1)))
        
        doc.build(story)
        buffer.seek(0)
        return buffer
